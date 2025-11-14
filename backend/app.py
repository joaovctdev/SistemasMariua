from flask import Flask, request, jsonify
from flask_cors import CORS
from datetime import datetime, timedelta
import jwt
import hashlib
import pandas as pd
import os
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import time
import shutil

app = Flask(__name__)
CORS(app)

# Configuração
app.config['SECRET_KEY'] = 'sua-chave-secreta-super-segura'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}

# Criar pasta de uploads se não existir
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Criar pasta para programação do dia
PROGRAMACAO_DIA_FOLDER = os.path.join(app.config['UPLOAD_FOLDER'], 'ProgramacaoNovembro')
if not os.path.exists(PROGRAMACAO_DIA_FOLDER):
    os.makedirs(PROGRAMACAO_DIA_FOLDER)

# Variável global para armazenar a programação do dia atual
programacao_dia_data = []

def allowed_file(filename):
    """
    Valida se o arquivo enviado possui extensão permitida (.xlsx ou .xls)

    Args:
        filename (str): Nome do arquivo a ser verificado

    Returns:
        bool: True se a extensão for permitida, False caso contrário

    Funcionalidade:
        - Verifica se o nome do arquivo contém um ponto (.)
        - Extrai a extensão após o último ponto
        - Compara com as extensões permitidas definidas em ALLOWED_EXTENSIONS
        - Usado para garantir que apenas planilhas Excel sejam aceitas no upload
    """
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def processar_programacao_dia(caminho_arquivo):
    """
    Processa arquivo Excel da programação diária de obras e retorna estrutura organizada

    Args:
        caminho_arquivo (str): Caminho completo do arquivo Excel a ser processado

    Returns:
        list: Lista de dicionários com os dados da programação diária

    Estrutura do arquivo Excel esperada (8 colunas):
        - Coluna 0: Data da programação
        - Coluna 1: Código do Projeto (ex: B-1234567)
        - Coluna 2: Nome do Supervisor responsável
        - Coluna 3: Nome do Encarregado da equipe
        - Coluna 4: Título/descrição da obra
        - Coluna 5: Município onde a obra está localizada
        - Coluna 6: Atividade programada para o dia (Locação, Implantação, etc)
        - Coluna 7: Critério da obra (QLP, QLU, etc)

    Funcionalidade:
        - Lê arquivo Excel e converte para DataFrame pandas
        - Itera por cada linha verificando se há dados válidos
        - Pula linhas vazias (onde projeto está vazio)
        - Cria objeto JSON para cada linha com estrutura padronizada
        - Retorna lista completa para ser utilizada pela API
        - Usado no endpoint /api/programacao-dia/upload
    """
    try:
        # Ler Excel usando pandas
        df = pd.read_excel(caminho_arquivo, header=0)

        print(f"📊 Colunas encontradas: {df.columns.tolist()}")
        print(f"📊 Total de linhas: {len(df)}")

        programacao = []

        for idx, row in df.iterrows():
            # Verificar se a linha tem dados válidos (pelo menos projeto preenchido)
            if pd.isna(row.iloc[1]) or str(row.iloc[1]).strip() == '':
                continue  # Pular linhas vazias

            # Estrutura: Data, Projeto, Supervisor, Encarregado, Título, Município, Atividade Programada, Critério
            item = {
                'id': idx + 1,
                'data': str(row.iloc[0]) if pd.notna(row.iloc[0]) else '',
                'projeto': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
                'supervisor': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
                'encarregado': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
                'titulo': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
                'municipio': str(row.iloc[5]) if pd.notna(row.iloc[5]) else '',
                'atividadeProgramada': str(row.iloc[6]) if pd.notna(row.iloc[6]) else '',
                'criterio': str(row.iloc[7]) if pd.notna(row.iloc[7]) else ''
            }

            programacao.append(item)

        print(f"✅ Total de itens válidos processados: {len(programacao)}")
        return programacao

    except Exception as e:
        print(f"❌ Erro ao processar programação do dia: {str(e)}")
        raise

# Banco de dados simulado
users_db = {
    'admin@mariua.net': {
        'password': hashlib.sha256('MARIUA2025'.encode()).hexdigest(),
        'name': 'Administrador'
    }
}

# Configurações padrão por projeto B
PROJETOS_CONFIG = {
    'B-0001': {
        'supervisor': 'João Silva',
        'cliente': 'Cliente Exemplo 1',
        'localidade': 'Irecê',
        'criterio': 'QLP'
    },
    'B-0002': {
        'supervisor': 'Maria Santos',
        'cliente': 'Cliente Exemplo 2',
        'localidade': 'Central',
        'criterio': 'QLU'
    }
}

# Lista de atividades disponíveis
ATIVIDADES_DISPONIVEIS = [
    'IMPLANTAÇÃO',
    'LANÇAMENTO',
    'ESCAVAÇÃO',
    'DESCARGA/IMPLANTAÇÃO',
    'INSTALAÇÃO DE ESTRUTURA',
    'PODA DE LIVRAMENTO',
    'CAVA COM RETRO'
]

# Função helper para salvar planilha com retry
def salvar_planilha_com_retry(wb, caminho, max_tentativas=3):
    """Tenta salvar a planilha com retry em caso de erro de permissão"""
    for tentativa in range(max_tentativas):
        try:
            # Criar backup temporário
            backup_path = caminho + '.backup'
            if os.path.exists(backup_path):
                os.remove(backup_path)
            
            # Salvar no arquivo temporário primeiro
            temp_path = caminho + '.temp'
            wb.save(temp_path)
            
            # Se salvou com sucesso, fazer backup do original
            if os.path.exists(caminho):
                shutil.copy2(caminho, backup_path)
            
            # Substituir o arquivo original
            shutil.move(temp_path, caminho)
            
            # Remover backup se tudo deu certo
            if os.path.exists(backup_path):
                os.remove(backup_path)
            
            print(f"✅ Planilha salva com sucesso!")
            return True
            
        except PermissionError as e:
            print(f"⚠️ Tentativa {tentativa + 1}/{max_tentativas} falhou: Arquivo pode estar aberto")
            if tentativa < max_tentativas - 1:
                time.sleep(1)  # Espera 1 segundo antes de tentar novamente
            else:
                raise Exception(f"❌ Não foi possível salvar o arquivo após {max_tentativas} tentativas. "
                              f"FECHE O ARQUIVO EXCEL se estiver aberto e tente novamente!")
        except Exception as e:
            print(f"❌ Erro ao salvar: {str(e)}")
            raise

    return False

# Função auxiliar para processar planilha
def processar_planilha(caminho_arquivo):
    """Processa a planilha e retorna lista de obras"""
    try:
        # Ler o Excel com pandas
        df = pd.read_excel(caminho_arquivo, header=0)
        
        # Processar os dados
        obras = []
        
        for index, row in df.iterrows():
            try:
                # Extrair TODAS as colunas (baseado na ordem correta do Excel)
                # A - ENCARREGADO
                encarregado = str(row.iloc[0]) if pd.notna(row.iloc[0]) else 'N/A'
                # B - SUPERVISOR
                supervisor = str(row.iloc[1]) if pd.notna(row.iloc[1]) else 'N/A'
                # C - PROJETO
                projeto = str(row.iloc[2]) if pd.notna(row.iloc[2]) else f"B-{str(index+1).zfill(4)}"
                # D - TÍTULO (cliente)
                cliente = str(row.iloc[3]) if pd.notna(row.iloc[3]) else 'N/A'
                # E - MUNICIPIO (localidade)
                localidade = str(row.iloc[4]) if pd.notna(row.iloc[4]) else 'N/A'
                # F - CRITÉRIO
                criterio = str(row.iloc[5]) if pd.notna(row.iloc[5]) else ''
                # G - ANOTAÇÕES
                anotacoes = str(row.iloc[6]) if pd.notna(row.iloc[6]) else ''
                # H - POSTES PREVISTOS
                postes_previstos = float(row.iloc[7]) if pd.notna(row.iloc[7]) else 0
                # I - DATA DE INÍCIO
                data_inicio = str(row.iloc[8]) if pd.notna(row.iloc[8]) else ''
                # J - DATA CONCLUSÃO (prazo)
                prazo = str(row.iloc[9]) if pd.notna(row.iloc[9]) else ''
                # K - OBRA DA SEMANA
                obra_semana = str(row.iloc[10]) if pd.notna(row.iloc[10]) else ''
                # L - MOTIVO DO ATRASO (NOVA COLUNA)
                motivo_atraso = str(row.iloc[11]) if pd.notna(row.iloc[11]) else ''
                # M - NECESSIDADE (ATIVIDADE DO DIA)
                atividade_dia = str(row.iloc[12]) if pd.notna(row.iloc[12]) else 'IMPLANTAÇÃO'
                # N - PROGRAMAÇÃO LV
                programacao_lv = str(row.iloc[13]) if pd.notna(row.iloc[13]) else ''
                # O - CAVAS REALIZADAS
                cavas_realizadas = float(row.iloc[14]) if pd.notna(row.iloc[14]) else 0
                # P - POSTES REALIZADOS
                postes_implantados = float(row.iloc[15]) if pd.notna(row.iloc[15]) else 0
                # Q - LATITUDE
                latitude_raw = row.iloc[16] if pd.notna(row.iloc[16]) else None
                latitude = None
                if latitude_raw is not None:
                    try:
                        # Tentar converter para float
                        lat_str = str(latitude_raw).replace(',', '.').strip()
                        latitude = float(lat_str)
                    except:
                        latitude = None
                
                # R - LONGITUDE
                longitude_raw = row.iloc[17] if pd.notna(row.iloc[17]) else None
                longitude = None
                if longitude_raw is not None:
                    try:
                        # Tentar converter para float
                        lng_str = str(longitude_raw).replace(',', '.').strip()
                        longitude = float(lng_str)
                    except:
                        longitude = None
                # S - CLIENTES PREVISTOS
                clientes_previstos = float(row.iloc[18]) if pd.notna(row.iloc[18]) else 0
                # T - PROJETO KIT
                projeto_kit = str(row.iloc[19]) if pd.notna(row.iloc[19]) else ''
                # U - PROJETO MEDIDOR
                projeto_medidor = str(row.iloc[20]) if pd.notna(row.iloc[20]) else ''
                # V - AR COELBA
                ar_coelba = str(row.iloc[21]) if pd.notna(row.iloc[21]) else 'N/A'
                # W - VISITA PRÉVIA
                data_visita_previa = str(row.iloc[22]) if pd.notna(row.iloc[22]) else ''
                # X - OBSERVAÇÃO DA VISITA
                observacao_visita = str(row.iloc[23]) if pd.notna(row.iloc[23]) else ''
                # Y - ANÁLISE PRÉ FECH
                analise_pre_fechamento = str(row.iloc[24]) if pd.notna(row.iloc[24]) else ''
                # Z - SOLICITAÇÃO DE RESERVA
                data_solicitacao_reserva = str(row.iloc[25]) if pd.notna(row.iloc[25]) else ''
                
                # Calcular progresso
                progresso = 0
                if postes_previstos > 0:
                    progresso = min(round((postes_implantados / postes_previstos) * 100), 100)
                
                # Verificar se está energizada
                is_energizada = 'ENERGIZADA' in anotacoes.upper()
                
                # Determinar status baseado nas DATAS (LÓGICA CORRIGIDA)
                from datetime import datetime
                hoje = datetime.now().date()
                
                # Processar data de início
                dt_inicio = None
                if data_inicio and data_inicio != 'nan' and str(data_inicio).strip():
                    try:
                        # Tentar converter data
                        if isinstance(data_inicio, datetime):
                            dt_inicio = data_inicio.date()
                        else:
                            # Tentar parsear string
                            data_str = str(data_inicio).strip()
                            if '/' in data_str:
                                dt_inicio = datetime.strptime(data_str, '%d/%m/%Y').date()
                            elif '-' in data_str:
                                dt_inicio = datetime.strptime(data_str.split()[0], '%Y-%m-%d').date()
                    except:
                        dt_inicio = None
                
                # Processar data de término (prazo)
                dt_termino = None
                if prazo and prazo != 'nan' and str(prazo).strip():
                    try:
                        if isinstance(prazo, datetime):
                            dt_termino = prazo.date()
                        else:
                            data_str = str(prazo).strip()
                            if '/' in data_str:
                                dt_termino = datetime.strptime(data_str, '%d/%m/%Y').date()
                            elif '-' in data_str:
                                dt_termino = datetime.strptime(data_str.split()[0], '%Y-%m-%d').date()
                    except:
                        dt_termino = None
                
                # ========== LÓGICA DE STATUS CORRIGIDA ==========
                # 1. ENERGIZADA: Anotações contém "ENERGIZADA" OU Data de término < hoje
                # 2. PROGRAMADA: Data início > hoje
                # 3. EM ANDAMENTO: Data início <= hoje E Data término >= hoje
                # 4. CONCLUÍDA: Progresso >= 100%
                
                if is_energizada or (dt_termino and dt_termino < hoje):
                    # Obras energizadas: anotações com "ENERGIZADA" ou data término passou
                    status = 'Energizada'
                elif progresso >= 100:
                    # Obras concluídas: progresso completo
                    status = 'Concluída'
                elif dt_inicio and dt_inicio > hoje:
                    # Obras programadas: data de início no futuro
                    status = 'Programada'
                elif dt_inicio and dt_inicio <= hoje:
                    # Obras que já começaram
                    if dt_termino:
                        if dt_termino >= hoje:
                            # Data início passou e data término não passou = Em Andamento
                            status = 'Em Andamento'
                        else:
                            # Data início passou e data término passou = Energizada
                            status = 'Energizada'
                    else:
                        # Sem data término definida, mas já começou = Em Andamento
                        status = 'Em Andamento'
                else:
                    # Se não tem data de início, usa progresso
                    if progresso > 0:
                        status = 'Em Andamento'
                    else:
                        status = 'Programada'
                
                # Verificar se tem coordenadas válidas
                has_valid_coordinates = (
                    latitude is not None and 
                    longitude is not None and
                    -90 <= latitude <= 90 and
                    -180 <= longitude <= 180
                )
                
                obra = {
                    'id': index + 1,
                    'encarregado': encarregado,
                    'supervisor': supervisor,
                    'projeto': projeto,
                    'cliente': cliente,
                    'localidade': localidade,
                    'criterio': criterio,
                    'anotacoes': anotacoes,
                    'postesPrevistos': int(postes_previstos),
                    'dataInicio': data_inicio,
                    'prazo': prazo,
                    'obraSemana': obra_semana,
                    'motivoAtraso': motivo_atraso,
                    'atividadeDia': atividade_dia,
                    'programacaoLv': programacao_lv,
                    'cavasRealizadas': int(cavas_realizadas),
                    'postesImplantados': int(postes_implantados),
                    'latitude': latitude,
                    'longitude': longitude,
                    'hasCoordinates': has_valid_coordinates,
                    'clientesPrevistos': int(clientes_previstos),
                    'projetoKit': projeto_kit,
                    'projetoMedidor': projeto_medidor,
                    'arCoelba': ar_coelba,
                    'dataVisitaPrevia': data_visita_previa,
                    'observacaoVisita': observacao_visita,
                    'analisePreFechamento': analise_pre_fechamento,
                    'dataSolicitacaoReserva': data_solicitacao_reserva,
                    'progresso': progresso,
                    'isEnergizada': is_energizada,
                    'status': status
                }
                
                # Só adicionar se tiver projeto válido
                if projeto and projeto != 'nan':
                    obras.append(obra)
                    # Debug: mostrar coordenadas das primeiras 3 obras
                    if len(obras) <= 3:
                        print(f"Obra {projeto}: Status={status}, DataInicio={dt_inicio}, DataTermino={dt_termino}, Hoje={hoje}")
                    
                    
            except Exception as e:
                print(f"Erro ao processar linha {index}: {e}")
                continue
        
        return obras
    except Exception as e:
        raise Exception(f"Erro ao processar planilha: {str(e)}")

# Endpoint de login
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': 'Usuário e senha são obrigatórios'}), 400
    
    user = users_db.get(username)
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    
    if user and user['password'] == password_hash:
        token = jwt.encode({
            'username': username,
            'exp': datetime.utcnow() + timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        
        return jsonify({
            'token': token,
            'username': username,
            'name': user['name']
        }), 200
    
    return jsonify({'error': 'Credenciais inválidas'}), 401

# Endpoint para upload de programação do dia
@app.route('/api/programacao-dia/upload', methods=['POST', 'OPTIONS'])
def upload_programacao_dia():
    global programacao_dia_data

    # Handle CORS preflight
    if request.method == 'OPTIONS':
        return '', 200

    try:
        print("📤 Recebendo upload de programação do dia...")
        print(f"📋 Files na request: {list(request.files.keys())}")

        # Verificar se há arquivo na requisição
        if 'file' not in request.files:
            print("❌ Nenhum arquivo na request")
            return jsonify({'error': 'Nenhum arquivo enviado', 'success': False}), 400

        file = request.files['file']
        print(f"📄 Arquivo recebido: {file.filename}")

        # Verificar se o arquivo tem nome
        if file.filename == '':
            print("❌ Nome de arquivo vazio")
            return jsonify({'error': 'Nome de arquivo vazio', 'success': False}), 400

        # Verificar extensão do arquivo
        if not allowed_file(file.filename):
            print(f"❌ Extensão não permitida: {file.filename}")
            return jsonify({
                'error': 'Tipo de arquivo não permitido. Apenas arquivos .xlsx ou .xls são aceitos',
                'success': False
            }), 400

        # Salvar temporariamente para processar
        temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_programacao_dia.xlsx')
        file.save(temp_filepath)

        print(f"✅ Arquivo temporário salvo: {temp_filepath}")
        print(f"✅ Tamanho do arquivo: {os.path.getsize(temp_filepath)} bytes")

        # Processar o arquivo
        try:
            programacao_dia_data = processar_programacao_dia(temp_filepath)
            total_itens = len(programacao_dia_data)
            print(f"✅ Programação processada: {total_itens} itens encontrados")

            # Remover arquivo temporário
            os.remove(temp_filepath)

            return jsonify({
                'success': True,
                'message': f'Programação do dia carregada com sucesso!',
                'total_itens': total_itens,
                'programacao': programacao_dia_data
            }), 200

        except Exception as e:
            # Remover arquivo temporário em caso de erro
            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)
            print(f"❌ Erro ao processar programação: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'error': f'Erro ao processar arquivo: {str(e)}. Verifique se as colunas estão na ordem: Data, Projeto, Supervisor, Encarregado, Título, Município, Atividade Programada, Critério',
                'success': False
            }), 400

    except Exception as e:
        print(f"❌ Erro no upload: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': f'Erro no upload: {str(e)}',
            'success': False
        }), 500

# Endpoint para buscar programação do dia atual
@app.route('/api/programacao-dia', methods=['GET'])
def get_programacao_dia():
    global programacao_dia_data

    return jsonify({
        'success': True,
        'total': len(programacao_dia_data),
        'programacao': programacao_dia_data
    }), 200

# Endpoint para salvar programação do dia na pasta ProgramacaoNovembro
@app.route('/api/programacao-dia/salvar', methods=['POST'])
def salvar_programacao_dia():
    global programacao_dia_data

    try:
        if not programacao_dia_data:
            return jsonify({
                'error': 'Nenhuma programação carregada. Faça upload de um arquivo primeiro.',
                'success': False
            }), 400

        # Obter dados atualizados do request (caso tenha modificações)
        data = request.get_json()
        if data and 'programacao' in data:
            programacao_dia_data = data['programacao']

        # Nome do arquivo com data do dia
        data_hoje = datetime.now().strftime('%d-%m-%Y')
        filename = f"{data_hoje}.xlsx"
        filepath = os.path.join(PROGRAMACAO_DIA_FOLDER, filename)

        # Criar DataFrame com as colunas na ordem especificada
        df = pd.DataFrame(programacao_dia_data)

        # Reordenar colunas
        colunas_ordenadas = ['data', 'projeto', 'supervisor', 'encarregado', 'titulo', 'municipio', 'atividadeProgramada', 'criterio']
        df = df[colunas_ordenadas]

        # Renomear colunas para o Excel
        df.columns = ['Data', 'Projeto', 'Supervisor', 'Encarregado', 'Título', 'Município', 'Atividade Programada', 'Critério']

        # Salvar no Excel
        df.to_excel(filepath, index=False, engine='openpyxl')

        print(f"✅ Programação salva: {filepath}")

        return jsonify({
            'success': True,
            'message': f'Programação salva com sucesso!',
            'filename': filename,
            'filepath': filepath
        }), 200

    except Exception as e:
        print(f"❌ Erro ao salvar programação: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': f'Erro ao salvar programação: {str(e)}',
            'success': False
        }), 500

# Endpoint para buscar obras do mês (arquivo fixo)
@app.route('/api/obras', methods=['GET'])
@app.route('/api/obras/', methods=['GET'])
def get_obras():
    try:
        # Arquivo fixo para obras do mês
        planilha_path = os.path.join(app.config['UPLOAD_FOLDER'], 'PROGRAMACAO - NOVEMBRO.xlsx')

        print(f"Procurando planilha em: {planilha_path}")
        print(f"Arquivo existe? {os.path.exists(planilha_path)}")

        if not os.path.exists(planilha_path):
            return jsonify({'error': 'Planilha PROGRAMACAO - NOVEMBRO.xlsx não encontrada no servidor'}), 404

        obras = processar_planilha(planilha_path)

        print(f"Total de obras processadas: {len(obras)}")

        return jsonify({
            'success': True,
            'total': len(obras),
            'obras': obras
        }), 200

    except Exception as e:
        print(f"Erro ao buscar obras: {e}")
        return jsonify({'error': str(e)}), 500

# Endpoint para buscar atividades disponíveis
@app.route('/api/atividades', methods=['GET'])
def get_atividades():
    return jsonify({
        'success': True,
        'atividades': ATIVIDADES_DISPONIVEIS
    }), 200

# Endpoint para buscar configuração de projeto B
@app.route('/api/projeto-config/<projeto_id>', methods=['GET'])
def get_projeto_config(projeto_id):
    config = PROJETOS_CONFIG.get(projeto_id, {})
    return jsonify({
        'success': True,
        'config': config
    }), 200

# Endpoint para adicionar nova obra
@app.route('/api/obras/adicionar', methods=['POST'])
def adicionar_obra():
    try:
        data = request.get_json()
        print(f"📥 Dados recebidos para adicionar: {data}")

        # Arquivo fixo para obras do mês
        planilha_path = os.path.join(app.config['UPLOAD_FOLDER'], 'PROGRAMACAO - NOVEMBRO.xlsx')

        if not os.path.exists(planilha_path):
            print("❌ ERRO: Planilha não encontrada")
            return jsonify({'error': 'Planilha PROGRAMACAO - NOVEMBRO.xlsx não encontrada'}), 404
        
        # Verificar se o arquivo está aberto
        try:
            # Tenta abrir o arquivo para verificar permissões
            with open(planilha_path, 'r+b'):
                pass
        except PermissionError:
            return jsonify({
                'error': '⚠️ O arquivo Excel está ABERTO! Por favor, FECHE o arquivo "PROGRAMACAO - NOVEMBRO.xlsx" e tente novamente.'
            }), 423  # 423 Locked
        
        # Carregar a planilha
        wb = load_workbook(planilha_path)
        ws = wb.active
        
        # Encontrar a próxima linha vazia
        proxima_linha = ws.max_row + 1
        print(f"➕ Adicionando obra na linha: {proxima_linha}")
        
        # Adicionar dados na planilha
        ws[f'A{proxima_linha}'] = data.get('encarregado', 'N/A')
        ws[f'B{proxima_linha}'] = data.get('supervisor', 'N/A')
        ws[f'C{proxima_linha}'] = data.get('projeto', '')
        ws[f'D{proxima_linha}'] = data.get('cliente', 'N/A')
        ws[f'E{proxima_linha}'] = data.get('localidade', 'N/A')
        ws[f'F{proxima_linha}'] = data.get('criterio', '')
        ws[f'G{proxima_linha}'] = data.get('anotacoes', '')
        ws[f'H{proxima_linha}'] = data.get('postesPrevistos', 0)
        ws[f'I{proxima_linha}'] = data.get('dataInicio', '')
        ws[f'J{proxima_linha}'] = data.get('prazo', '')
        ws[f'K{proxima_linha}'] = data.get('obraSemana', '')
        ws[f'L{proxima_linha}'] = data.get('atividadeDia', 'IMPLANTAÇÃO')
        ws[f'M{proxima_linha}'] = data.get('programacaoLv', '')
        ws[f'N{proxima_linha}'] = data.get('cavasRealizadas', 0)
        ws[f'O{proxima_linha}'] = data.get('postesImplantados', 0)
        ws[f'P{proxima_linha}'] = data.get('latitude', '')
        ws[f'Q{proxima_linha}'] = data.get('longitude', '')
        ws[f'R{proxima_linha}'] = data.get('clientesPrevistos', 0)
        ws[f'S{proxima_linha}'] = data.get('projetoKit', '')
        ws[f'T{proxima_linha}'] = data.get('projetoMedidor', '')
        ws[f'U{proxima_linha}'] = data.get('arCoelba', 'N/A')
        ws[f'V{proxima_linha}'] = data.get('dataVisitaPrevia', '')
        ws[f'W{proxima_linha}'] = data.get('observacaoVisita', '')
        ws[f'X{proxima_linha}'] = data.get('analisePreFechamento', '')
        ws[f'Y{proxima_linha}'] = data.get('dataSolicitacaoReserva', '')
        
        print(f"💾 Salvando planilha...")
        
        # Salvar a planilha com retry
        try:
            salvar_planilha_com_retry(wb, planilha_path)
            wb.close()
            
            print("✅ Obra adicionada com sucesso!")
            return jsonify({
                'success': True,
                'message': 'Obra adicionada com sucesso!'
            }), 200
            
        except Exception as save_error:
            wb.close()
            raise save_error
        
    except Exception as e:
        print(f"❌ ERRO ao adicionar obra: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Mensagem mais amigável para erro de permissão
        error_msg = str(e)
        if 'Permission denied' in error_msg or 'FECHE O ARQUIVO' in error_msg:
            error_msg = '⚠️ O arquivo Excel está ABERTO! Por favor, FECHE o arquivo e tente novamente.'
        
        return jsonify({'error': error_msg}), 500

# Endpoint para atualizar obra existente
@app.route('/api/obras/atualizar/<int:obra_id>', methods=['PUT'])
def atualizar_obra(obra_id):
    try:
        data = request.get_json()

        # Arquivo fixo para obras do mês
        planilha_path = os.path.join(app.config['UPLOAD_FOLDER'], 'PROGRAMACAO - NOVEMBRO.xlsx')

        print(f"📝 Tentando atualizar obra ID: {obra_id}")
        print(f"📥 Dados recebidos: {data}")

        if not os.path.exists(planilha_path):
            return jsonify({'error': 'Planilha PROGRAMACAO - NOVEMBRO.xlsx não encontrada'}), 404
        
        # Verificar se o arquivo está aberto
        try:
            with open(planilha_path, 'r+b'):
                pass
        except PermissionError:
            return jsonify({
                'error': '⚠️ O arquivo Excel está ABERTO! Por favor, FECHE o arquivo "PROGRAMACAO - NOVEMBRO.xlsx" e tente novamente.'
            }), 423  # 423 Locked
        
        # Carregar a planilha
        wb = load_workbook(planilha_path)
        ws = wb.active
        
        # A linha na planilha é obra_id (obra_id já considera o cabeçalho)
        # obra_id = 1 significa a primeira obra (linha 2 no Excel, pois linha 1 é cabeçalho)
        linha = obra_id + 1
        
        print(f"🔄 Atualizando linha {linha} na planilha")
        
        # Verificar se a linha existe
        if linha > ws.max_row:
            wb.close()
            return jsonify({'error': f'Obra ID {obra_id} não encontrada'}), 404
        
        # Atualizar dados na planilha (apenas campos fornecidos)
        if 'encarregado' in data:
            ws[f'A{linha}'] = data['encarregado']
            print(f"✏️ Encarregado atualizado: {data['encarregado']}")
        
        if 'supervisor' in data:
            ws[f'B{linha}'] = data['supervisor']
            print(f"✏️ Supervisor atualizado: {data['supervisor']}")
        
        if 'atividadeDia' in data:
            ws[f'L{linha}'] = data['atividadeDia']
            print(f"✏️ Atividade do dia atualizada: {data['atividadeDia']}")
        
        if 'postesImplantados' in data:
            ws[f'O{linha}'] = data['postesImplantados']
            print(f"✏️ Postes implantados atualizado: {data['postesImplantados']}")
        
        if 'cavasRealizadas' in data:
            ws[f'N{linha}'] = data['cavasRealizadas']
            print(f"✏️ Cavas realizadas atualizado: {data['cavasRealizadas']}")
        
        if 'anotacoes' in data:
            ws[f'G{linha}'] = data['anotacoes']
            print(f"✏️ Anotações atualizadas: {data['anotacoes']}")
        
        print(f"💾 Salvando alterações...")
        
        # Salvar a planilha com retry
        try:
            salvar_planilha_com_retry(wb, planilha_path)
            wb.close()
            
            print(f"✅ Obra atualizada com sucesso!")
            return jsonify({
                'success': True,
                'message': 'Obra atualizada com sucesso'
            }), 200
            
        except Exception as save_error:
            wb.close()
            raise save_error
        
    except Exception as e:
        print(f"❌ ERRO ao atualizar obra: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Mensagem mais amigável para erro de permissão
        error_msg = str(e)
        if 'Permission denied' in error_msg or 'FECHE O ARQUIVO' in error_msg:
            error_msg = '⚠️ O arquivo Excel está ABERTO! Por favor, FECHE o arquivo e tente novamente.'
        
        return jsonify({'error': error_msg}), 500

# Rota de teste
@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'}), 200

# Endpoint de debug para coordenadas
@app.route('/api/debug/coordenadas', methods=['GET'])
def debug_coordenadas():
    try:
        # Arquivo fixo para obras do mês
        planilha_path = os.path.join(app.config['UPLOAD_FOLDER'], 'PROGRAMACAO - NOVEMBRO.xlsx')

        if not os.path.exists(planilha_path):
            return jsonify({'error': 'Planilha PROGRAMACAO - NOVEMBRO.xlsx não encontrada'}), 404
        
        df = pd.read_excel(planilha_path, header=0)
        
        coordenadas_info = []
        for index, row in df.iterrows():
            if index < 5:  # Apenas as 5 primeiras linhas
                projeto = str(row.iloc[2]) if pd.notna(row.iloc[2]) else 'N/A'
                lat_raw = row.iloc[15]
                lng_raw = row.iloc[16]
                
                coordenadas_info.append({
                    'linha': index + 1,
                    'projeto': projeto,
                    'latitude_raw': str(lat_raw),
                    'longitude_raw': str(lng_raw),
                    'lat_type': str(type(lat_raw)),
                    'lng_type': str(type(lng_raw))
                })
        
        return jsonify({
            'total_linhas': len(df),
            'primeiras_5_linhas': coordenadas_info
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/bd-programacao', methods=['GET'])
def get_bd_programacao():
    """Retorna dados do banco de dados MainBD para dashboards"""
    try:
        caminho_bd = os.path.join(app.config['UPLOAD_FOLDER'], 'BD', 'MainBD.xlsx')

        if not os.path.exists(caminho_bd):
            return jsonify({'error': 'Arquivo MainBD.xlsx não encontrado'}), 404

        # Ler a planilha
        df = pd.read_excel(caminho_bd, engine='openpyxl')

        # Processar os dados conforme a estrutura do MainBD
        dados = []
        for index, row in df.iterrows():
            try:
                # Mapeamento das colunas do MainBD:
                # 1 = data_servico, 5 = des_equipe (encarregado), 7 = Supervisor, 
                # 2 = Contrato (projeto), 21 = localidade (municipio),
                # 31 = des_atividade (atividade), 34 = qtd_atividade (quantidade),
                # 32 = latitude, 33 = longitude, 35 = valor_unitario
                
                des_atividade = str(row.iloc[31]).upper() if pd.notna(row.iloc[31]) else ''
                qtd_atividade = float(row.iloc[34]) if pd.notna(row.iloc[34]) else 0
                
                # Determinar tipo de atividade baseado na descrição
                poste_real = 0
                cava_real = 0
                
                # POSTE: tudo que contém "POSTE BT" ou "POSTE AT"
                if 'POSTE BT' in des_atividade or 'POSTE AT' in des_atividade:
                    poste_real = qtd_atividade
                
                # CAVA: tudo que contém "ESCAVAÇÃO" ou "CAVA"
                elif ('ESCAVAÇÃO' in des_atividade or 'ESCAVACAO' in des_atividade or 'CAVA' in des_atividade) and \
                    'ESTAI' not in des_atividade.upper() and 'ESTAÍ' not in des_atividade.upper():
                    cava_real = qtd_atividade
                
                registro = {
                    'data': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
                    'supervisor': str(row.iloc[6]) if pd.notna(row.iloc[6]) else '',
                    'encarregado': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
                    'projeto': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
                    'municipio': str(row.iloc[21]) if pd.notna(row.iloc[21]) else '',
                    'atividade': des_atividade,
                    'quantidade': qtd_atividade,
                    'posteReal': poste_real,
                    'cavaReal': cava_real,
                    'latitude': float(row.iloc[32]) if pd.notna(row.iloc[32]) else 0,
                    'longitude': float(row.iloc[33]) if pd.notna(row.iloc[33]) else 0,
                    'valorUnitario': float(row.iloc[35]) if pd.notna(row.iloc[35]) else 0,
                }
                dados.append(registro)
            except Exception as e:
                print(f"Erro ao processar linha {index}: {str(e)}")
                continue

        return jsonify({
            'success': True,
            'dados': dados,
            'total': len(dados)
        })

    except Exception as e:
        print(f"Erro ao carregar MainBD: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/obras-programacao', methods=['GET'])
def get_obras_programacao():
    """Retorna dados da planilha PROGRAMAÇÃO (postes previstos por obra)"""
    try:
        caminho_programacao = os.path.join(app.config['UPLOAD_FOLDER'], 'PROGRAMACAO - NOVEMBRO.xlsx')

        if not os.path.exists(caminho_programacao):
            return jsonify({'error': 'Arquivo PROGRAMACAO não encontrado'}), 404

        df = pd.read_excel(caminho_programacao, engine='openpyxl')

        obras = []
        for index, row in df.iterrows():
            if index == 0:  # Pular cabeçalho
                continue

            try:
                obra = {
                    'encarregado': str(row.iloc[0]) if pd.notna(row.iloc[0]) else '',
                    'supervisor': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
                    'projeto': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
                    'titulo': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
                    'municipio': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
                    'postesPrevisto': float(row.iloc[7]) if pd.notna(row.iloc[7]) and str(row.iloc[7]).replace('.','').replace('-','').isdigit() else 0,
                }
                obras.append(obra)
            except Exception as e:
                print(f"Erro ao processar obra linha {index}: {str(e)}")
                continue

        return jsonify({
            'success': True,
            'obras': obras,
            'total': len(obras)
        })

    except Exception as e:
        print(f"Erro ao carregar obras programação: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/cavas-por-retro', methods=['GET'])
def get_cavas_por_retro():
    """
    Endpoint: GET /api/dashboard/cavas-por-retro

    Retorna dados de escavações (cavas) realizadas por retroescavadeira para equipes especializadas

    Query Parameters:
        - mes (opcional): Número do mês (1-12) para filtrar. Default: 'todos'
        - semana (opcional): Número da semana do mês (1-4/5) para filtrar. Default: 'todos'

    Fonte de Dados:
        - Arquivo: uploads/BD/MainBD.xlsx (página BD)
        - Usa coluna 'data_servico' para filtros temporais
        - Usa coluna 'des_equipe' para identificar equipes
        - Usa coluna 'des_atividade' para classificar tipos de cava
        - Usa coluna 'qtd_atividade' para quantificar cavas

    Equipes Monitoradas:
        - JOAO-JAC, MENEZES-IRC, TIAGO-JAC, VAGNO-IRC, WESLEY-IRC, OSIMAR-JAC

    Tipos de Cava Classificados:
        - Cava Normal: Escavação em solo comum
        - Cava com Rompedor: Escavação que requer rompedor hidráulico
        - Cava em Rocha: Escavação em terreno rochoso

    Retorna:
        JSON com:
        - dados: Lista de equipes com totalizações por tipo de cava
        - detalhes: Registros individuais de cada serviço
        - filtros_aplicados: Filtros que foram utilizados na consulta

    Usado em:
        - Dashboard principal para gráfico de "Cavas por Retro"
        - Mostra produtividade das equipes especializadas em escavação
    """
    try:
        from datetime import datetime, timedelta
        from flask import request

        # Obter filtros da query string (parâmetros da URL)
        filtro_mes = request.args.get('mes', 'todos')
        filtro_semana = request.args.get('semana', 'todos')

        caminho_bd = os.path.join(app.config['UPLOAD_FOLDER'], 'BD', 'MainBD.xlsx')

        if not os.path.exists(caminho_bd):
            return jsonify({'error': 'Arquivo MainBD.xlsx não encontrado'}), 404

        # Ler a planilha BD do MainBD
        df = pd.read_excel(caminho_bd, engine='openpyxl')

        # Converter data_servico para formato datetime para permitir filtros
        df['data_servico'] = pd.to_datetime(df['data_servico'], errors='coerce')

        # Aplicar filtro de mês (se selecionado)
        if filtro_mes != 'todos':
            mes_num = int(filtro_mes)
            df = df[df['data_servico'].dt.month == mes_num]

        # Aplicar filtro de semana (se selecionado e mês também foi selecionado)
        if filtro_semana and filtro_semana != 'todos' and filtro_semana.strip() != '' and filtro_mes != 'todos':
            semana_num = int(filtro_semana)
            mes_num = int(filtro_mes)

            # Calcular intervalo da semana baseado no primeiro dia do mês
            ano_atual = df['data_servico'].dt.year.mode()[0] if not df.empty else datetime.now().year
            primeiro_dia_mes = datetime(int(ano_atual), mes_num, 1)

            # Calcular início e fim da semana (7 dias cada semana)
            inicio_semana = primeiro_dia_mes + timedelta(days=(semana_num - 1) * 7)
            fim_semana = inicio_semana + timedelta(days=6)

            # Filtrar dados apenas do intervalo da semana
            df = df[(df['data_servico'] >= inicio_semana) & (df['data_servico'] <= fim_semana)]

        # Equipes especializadas em escavação com retroescavadeira
        equipes_retro = ['JOAO-JAC', 'MENEZES-IRC', 'TIAGO-JAC', 'VAGNO-IRC', 'WESLEY-IRC', 'OSIMAR-JAC']

        # Inicializar estrutura de dados para cada equipe
        dados_retro = {}
        for equipe in equipes_retro:
            dados_retro[equipe] = {
                'total_cavas': 0,
                'cavas_normal': 0,
                'cavas_rompedor': 0,
                'cavas_rocha': 0,
                'registros': []
            }

        # Processar cada linha do DataFrame
        for index, row in df.iterrows():
            try:
                # Obter nome da equipe e descrição da atividade
                des_equipe = str(row['des_equipe']).strip().upper() if pd.notna(row['des_equipe']) else ''
                des_atividade = str(row['des_atividade']).upper() if pd.notna(row['des_atividade']) else ''
                qtd_atividade = float(row['qtd_atividade']) if pd.notna(row['qtd_atividade']) else 0

                # Verificar se é uma equipe monitorada E se a atividade é de cava com retro
                if des_equipe in equipes_retro and 'RETRO' in des_atividade and 'CAVA' in des_atividade:
                    # Classificar tipo de cava pela descrição da atividade
                    if 'ROMPEDOR' in des_atividade:
                        dados_retro[des_equipe]['cavas_rompedor'] += qtd_atividade
                    elif 'ROCHA' in des_atividade:
                        dados_retro[des_equipe]['cavas_rocha'] += qtd_atividade
                    elif 'NORMAL' in des_atividade:
                        dados_retro[des_equipe]['cavas_normal'] += qtd_atividade

                    # Totalizar cavas
                    dados_retro[des_equipe]['total_cavas'] += qtd_atividade

                    # Armazenar registro detalhado para auditoria
                    dados_retro[des_equipe]['registros'].append({
                        'data': str(row['data_servico']) if pd.notna(row['data_servico']) else '',
                        'atividade': des_atividade,
                        'quantidade': qtd_atividade,
                        'localidade': str(row['bairro']) if pd.notna(row['bairro']) else '',
                        'supervisor': str(row['Supervisor']) if pd.notna(row['Supervisor']) else ''
                    })

            except Exception as e:
                print(f"Erro ao processar linha {index}: {str(e)}")
                continue

        # Preparar resposta com estatísticas agregadas
        resultado = []
        for equipe, dados in dados_retro.items():
            resultado.append({
                'equipe': equipe,
                'total_cavas': dados['total_cavas'],
                'cavas_normal': dados['cavas_normal'],
                'cavas_rompedor': dados['cavas_rompedor'],
                'cavas_rocha': dados['cavas_rocha'],
                'total_registros': len(dados['registros'])
            })

        return jsonify({
            'success': True,
            'dados': resultado,
            'detalhes': dados_retro,
            'total_equipes': len(equipes_retro),
            'filtros_aplicados': {
                'mes': filtro_mes,
                'semana': filtro_semana
            }
        })

    except Exception as e:
        print(f"Erro ao carregar cavas por retro: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/utd-dados', methods=['GET'])
def get_utd_dados():
    """
    Endpoint: GET /api/dashboard/utd-dados

    Retorna dados consolidados das Unidades de Trabalho Descentralizadas (UTDs) IRECÊ e JACOBINA

    Fonte de Dados:
        - Arquivo: uploads/BD/MainBD.xlsx
        - Páginas: 'UTDIRECE' e 'UTDJACOBINA'

    Estrutura do MainBD (21 colunas por página UTD):
        - pep_obra: Código único do projeto
        - titulo: Nome/descrição da obra (usado para contagem)
        - municipio: Cidade onde a obra está localizada
        - localidade: Bairro/região específica
        - status: Estado atual da obra (ENERGIZADA, EM EXECUÇÃO, etc)
        - encarregado: Responsável pela equipe
        - supervisor: Supervisor da obra
        - ar_coelba: Agente Regional da COELBA responsável
        - plan_inv: Planejamento de investimento
        - criterio: Critério técnico (QLP, QLU, etc)
        - clientes_prev: Número de clientes atendidos pela obra
        - postes_imp: Postes implantados
        - inicio_previsto: Data prevista de início
        - fim_previsto: Data prevista de conclusão
        - data_energização: Data real de energização
        - valor_projeto: Valor total do projeto em R$
        - latitude/longitude: Coordenadas geográficas

    Retorna:
        JSON com 3 regiões (irece, jacobina, geral):
        - total_obras: Contagem de obras pelo campo 'titulo'
        - obras_energizadas: Obras com status ENERGIZADA
        - clientes: Soma de clientes_prev
        - valor_total: Soma de valor_projeto (limpo de formatação brasileira)
        - obras_por_ar: Dicionário com contagem de obras por AR_COELBA

    Usado em:
        - Dashboard principal para cards de clientes e valor
        - Gráfico de obras energizadas por região
        - Gráfico de distribuição de obras por AR COELBA
    """
    try:
        caminho_mainbd = os.path.join(app.config['UPLOAD_FOLDER'], 'BD', 'MainBD.xlsx')

        if not os.path.exists(caminho_mainbd):
            return jsonify({'error': 'Arquivo MainBD.xlsx não encontrado'}), 404

        # Ler as duas páginas principais do MainBD (uma para cada UTD)
        df_irece = pd.read_excel(caminho_mainbd, sheet_name='UTDIRECE', engine='openpyxl')
        df_jacobina = pd.read_excel(caminho_mainbd, sheet_name='UTDJACOBINA', engine='openpyxl')

        def limpar_valor_moeda(valor):
            """
            Converte valores monetários do formato brasileiro para float

            Formatos aceitos:
                - Float/int direto: 1234.56
                - Formato BR: R$ 1.234,56
                - Formato BR sem símbolo: 1.234,56

            Processo:
                1. Remove "R$" e espaços
                2. Remove pontos (separadores de milhares)
                3. Substitui vírgula por ponto (separador decimal)
                4. Converte para float

            Exemplos:
                "R$ 116.999,07" -> 116999.07
                "20073.12" -> 20073.12
                None -> 0
            """
            if pd.isna(valor):
                return 0
            if isinstance(valor, (int, float)):
                return float(valor)
            # Remover "R$", espaços, pontos (milhares) e substituir vírgula por ponto (decimal)
            valor_str = str(valor).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
            try:
                return float(valor_str)
            except:
                return 0

        def processar_regiao(df):
            """
            Processa dados de uma região (IRECÊ ou JACOBINA) e retorna métricas agregadas

            Args:
                df (DataFrame): DataFrame com dados da região

            Returns:
                dict: Dicionário com métricas totalizadas:
                    - total_obras: Quantidade total de obras (contadas por titulo não vazio)
                    - obras_energizadas: Obras com status 'ENERGIZADA'
                    - clientes: Soma de clientes previstos
                    - valor_total: Soma total dos valores dos projetos em R$
                    - obras_por_ar: Quantidade de obras por cada AR_COELBA

            Lógica:
                - Conta obras pelo campo 'titulo' (mais confiável que contar linhas)
                - Filtra status que contenha 'ENERGIZADA' (case insensitive)
                - Converte clientes_prev para numérico (trata erros como 0)
                - Limpa e soma valor_projeto (trata formato BR)
                - Agrupa e conta obras por ar_coelba (exclui vazios)
            """
            # Contar obras pelo campo titulo (não vazio)
            total_obras = df['titulo'].notna().sum()

            # Contar obras energizadas (busca case-insensitive por ENERGIZADA no status)
            obras_energizadas = df[df['status'].str.upper().str.contains('ENERGIZADA', na=False)].shape[0]

            # Somar clientes (converter para numérico, erros viram 0)
            clientes = pd.to_numeric(df['clientes_prev'], errors='coerce').fillna(0).sum()

            # Somar valor_projeto (limpar formato brasileiro e converter)
            df_copy = df.copy()
            df_copy['valor_limpo'] = df_copy['valor_projeto'].apply(limpar_valor_moeda)
            valor_total = df_copy['valor_limpo'].sum()

            # Contar obras por AR_COELBA (filtrar valores vazios)
            obras_por_ar = df[df['ar_coelba'].notna() & (df['ar_coelba'] != '')].groupby('ar_coelba').size().to_dict()

            return {
                'total_obras': int(total_obras),
                'obras_energizadas': int(obras_energizadas),
                'clientes': int(clientes),
                'valor_total': round(float(valor_total), 2),
                'obras_por_ar': obras_por_ar
            }

        # Processar cada região
        dados_irece = processar_regiao(df_irece)
        dados_jacobina = processar_regiao(df_jacobina)

        # Combinar dados gerais
        df_geral = pd.concat([df_irece, df_jacobina], ignore_index=True)
        dados_geral = processar_regiao(df_geral)

        # Preparar resultado
        resultado = {
            'irece': dados_irece,
            'jacobina': dados_jacobina,
            'geral': dados_geral
        }

        return jsonify({
            'success': True,
            'dados': resultado
        })

    except Exception as e:
        print(f"Erro ao carregar dados UTD: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/producao-dia', methods=['GET'])
def get_producao_dia():
    """Retorna dados da produção do dia com previsão e realizado usando MainBD"""
    try:
        # Obter data do parâmetro ou usar data atual
        from datetime import datetime
        from flask import request
        data_param = request.args.get('data', None)

        if data_param:
            data_hoje = data_param  # Formato esperado: DD-MM-YYYY
        else:
            data_hoje = datetime.now().strftime('%d-%m-%Y')

        # Caminho para a planilha de programação do dia
        caminho_programacao_dia = os.path.join(app.config['UPLOAD_FOLDER'], 'ProgramacaoNovembro', f'{data_hoje}.xlsx')

        if not os.path.exists(caminho_programacao_dia):
            return jsonify({'error': f'Arquivo de programação para {data_hoje} não encontrado'}), 404

        # Ler programação do dia (colunas B, C, D, E, G)
        df_programacao = pd.read_excel(caminho_programacao_dia, engine='openpyxl')

        # Ler banco de dados MainBD
        caminho_bd = os.path.join(app.config['UPLOAD_FOLDER'], 'BD', 'MainBD.xlsx')

        if not os.path.exists(caminho_bd):
            return jsonify({'error': 'Arquivo MainBD.xlsx não encontrado'}), 404

        df_bd = pd.read_excel(caminho_bd, engine='openpyxl')

        # Processar dados
        producoes = []

        for index, row_prog in df_programacao.iterrows():
            try:
                # Dados da programação (previsto)
                projeto = str(row_prog.iloc[1]) if pd.notna(row_prog.iloc[1]) else ''  # Coluna B (índice 1)
                supervisor = str(row_prog.iloc[2]) if pd.notna(row_prog.iloc[2]) else ''  # Coluna C (índice 2)
                encarregado = str(row_prog.iloc[3]) if pd.notna(row_prog.iloc[3]) else ''  # Coluna D (índice 3)
                titulo = str(row_prog.iloc[4]) if pd.notna(row_prog.iloc[4]) else ''  # Coluna E (índice 4)
                atividade_programada = str(row_prog.iloc[6]) if pd.notna(row_prog.iloc[6]) else ''  # Coluna G (índice 6)

                if not projeto:  # Pular linhas vazias
                    continue

                # Buscar dados realizados no MainBD correspondentes ao projeto, data e atividade
                # Coluna 2 = Contrato (PROJETO), Coluna 1 = data_servico no MainBD
                projeto_limpo = str(projeto).strip() if projeto else ''
                atividade_limpa = str(atividade_programada).strip().upper() if atividade_programada else ''
                
                # Converter data do parâmetro para formato de comparação
                data_hoje_formatada = data_hoje  # Formato: DD-MM-YYYY
                # Extrair dia, mês e ano para comparação
                partes_data = data_hoje_formatada.split('-')
                if len(partes_data) == 3:
                    dia_busca = partes_data[0]
                    mes_busca = partes_data[1]
                    ano_busca = partes_data[2]
                else:
                    dia_busca = mes_busca = ano_busca = ''
                
                # Filtrar por projeto (Contrato)
                mask_projeto = df_bd.iloc[:, 2].astype(str).str.strip() == projeto_limpo
                
                # Comparar data de forma flexível (pode estar como DD-MM-YYYY, DD/MM/YYYY, etc)
                def match_data(data_str):
                    if pd.isna(data_str):
                        return False
                    data_str_clean = str(data_str).strip()
                    # Verificar se contém dia, mês e ano
                    return (dia_busca in data_str_clean and mes_busca in data_str_clean and ano_busca in data_str_clean)
                
                mask_data = df_bd.iloc[:, 1].apply(match_data)
                dados_realizados = df_bd[mask_projeto & mask_data]

                # Inicializar valores realizados
                locacao = ''
                cava_real = 0
                cava_em_rocha = 0
                poste_real = 0
                evento = ''
                responsavel = ''
                justificativa = ''
                progresso = 0
                status = 'Não Concluído'

                if not dados_realizados.empty:
                    # Processar todas as linhas do MainBD para este projeto/data
                    for _, row_bd in dados_realizados.iterrows():
                        try:
                            # Coluna 31 = des_atividade (descrição da atividade)
                            des_atividade = str(row_bd.iloc[31]).upper() if pd.notna(row_bd.iloc[31]) else ''
                            # Coluna 34 = qtd_atividade (quantidade)
                            qtd_atividade = float(row_bd.iloc[34]) if pd.notna(row_bd.iloc[34]) else 0
                            
                            # IDENTIFICAR TIPO DE ATIVIDADE BASEADO NA DESCRIÇÃO
                            # POSTE: tudo que contém "POSTE BT" ou "POSTE AT"
                            if 'POSTE BT' in des_atividade or 'POSTE AT' in des_atividade:
                                poste_real += qtd_atividade
                            
                            # CAVA: tudo que contém "ESCAVAÇÃO" ou "CAVA"
                            elif 'ESCAVAÇÃO' in des_atividade or 'ESCAVACAO' in des_atividade or 'CAVA' in des_atividade:
                                # Verificar se é cava em rocha (contém "ROCHA")
                                if 'ROCHA' in des_atividade:
                                    cava_em_rocha += qtd_atividade
                                else:
                                    cava_real += qtd_atividade
                            
                            # LOCAÇÃO: atividades de locação
                            elif 'LOCAÇÃO' in des_atividade or 'LOCACAO' in des_atividade:
                                locacao = str(qtd_atividade) if qtd_atividade > 0 else ''
                            
                            # Outras informações relevantes
                            # Coluna 25 = retorno_campo (pode ser usado como evento/justificativa)
                            retorno_campo = str(row_bd.iloc[25]) if pd.notna(row_bd.iloc[25]) else ''
                            if retorno_campo and not justificativa:
                                justificativa = retorno_campo
                            
                            # Coluna 26 = observacao
                            observacao = str(row_bd.iloc[26]) if pd.notna(row_bd.iloc[26]) else ''
                            if observacao and not justificativa:
                                justificativa = observacao
                                
                        except Exception as e:
                            print(f"Erro ao processar linha do MainBD: {str(e)}")
                            continue

                    # Coluna 5 = des_equipe (responsável)
                    responsavel = str(dados_realizados.iloc[0].iloc[5]) if pd.notna(dados_realizados.iloc[0].iloc[5]) else ''

                    # Calcular progresso e status com base na atividade programada
                    atividade_upper = atividade_programada.upper()

                    # 1. IMPLANTAÇÃO: se POSTE REAL > 0, progresso = 100%, status = "CONCLUÍDO"
                    if 'IMPLANTAÇÃO' in atividade_upper or 'IMPLANTACAO' in atividade_upper:
                        if poste_real > 0:
                            progresso = 100
                            status = 'Concluído'
                        else:
                            progresso = 0
                            status = 'Não concluído'

                    # 2. ESCAVAÇÃO: se CAVA REAL > 0 OU CAVA EM ROCHA > 0, progresso = 100%
                    elif 'ESCAVAÇÃO' in atividade_upper or 'ESCAVACAO' in atividade_upper:
                        if cava_real > 0 or cava_em_rocha > 0:
                            progresso = 100
                            status = 'Concluído'
                        else:
                            progresso = 0
                            status = 'Não concluído'

                    # 3. ENERGIZAÇÃO: se justificativa contém "ENERGIZADA" ou "EXECUTADO"
                    elif 'ENERGIZAÇÃO' in atividade_upper or 'ENERGIZACAO' in atividade_upper:
                        justificativa_upper = justificativa.upper() if justificativa else ''
                        if 'ENERGIZADA' in justificativa_upper or 'EXECUTADO' in justificativa_upper:
                            progresso = 100
                            status = 'Concluído'
                        else:
                            progresso = 0
                            status = 'Não concluído'

                    # 4. LOCAÇÃO: se LOCAÇÃO > 0, progresso = 100%
                    elif 'LOCAÇÃO' in atividade_upper or 'LOCACAO' in atividade_upper:
                        try:
                            if locacao and locacao != '' and locacao != 'nan' and float(locacao) > 0:
                                progresso = 100
                                status = 'Concluído'
                            else:
                                progresso = 0
                                status = 'Não Concluído'
                        except:
                            progresso = 0
                            status = 'Não Concluído'

                    # 5. LANÇAMENTO: se justificativa contém "Lançamento"
                    elif 'LANÇAMENTO' in atividade_upper or 'LANCAMENTO' in atividade_upper:
                        justificativa_upper = justificativa.upper() if justificativa else ''
                        if 'LANÇAMENTO' in justificativa_upper or 'LANCAMENTO' in justificativa_upper or 'LANÇOU' in justificativa_upper or 'LANCOU' in justificativa_upper:
                            progresso = 100
                            status = 'Concluído'
                        else:
                            progresso = 0
                            status = 'Não Concluído'

                    # 6. Outras atividades: usar lógica baseada em quantidades
                    else:
                        total_realizado = cava_real + cava_em_rocha + poste_real
                        if total_realizado > 0:
                            progresso = min(100, int((total_realizado / 10) * 100))
                            if progresso >= 80:
                                status = 'Concluído'
                            elif progresso >= 50:
                                status = 'Em Andamento'
                            elif progresso > 0:
                                status = 'Iniciado'
                            else:
                                status = 'Não concluído'
                        else:
                            progresso = 0
                            status = 'Não concluído'

                producao = {
                    'projeto': projeto,
                    'supervisor': supervisor,
                    'encarregado': encarregado,
                    'titulo': titulo,
                    'atividadeProgramada': atividade_programada,
                    'locacao': locacao,
                    'cavaReal': cava_real,
                    'cavaEmRocha': cava_em_rocha,
                    'posteReal': poste_real,
                    'evento': evento,
                    'responsavel': responsavel,
                    'justificativa': justificativa,
                    'progresso': progresso,
                    'status': status
                }

                producoes.append(producao)

            except Exception as e:
                print(f"Erro ao processar linha {index}: {str(e)}")
                continue

        return jsonify({
            'success': True,
            'data': data_hoje,
            'producoes': producoes,
            'total': len(producoes)
        })

    except Exception as e:
        print(f"Erro ao carregar produção do dia: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("Iniciando servidor Flask...")
    print(f"Diretorio de uploads: {app.config['UPLOAD_FOLDER']}")
    print("Para carregar dados, faca upload de um arquivo Excel via interface web")
    print("API rodando em: http://localhost:5000")
    app.run(debug=True, port=5000)